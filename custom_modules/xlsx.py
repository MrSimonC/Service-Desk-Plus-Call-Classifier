import csv
import warnings
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle
from collections import OrderedDict
from typing import Dict, List
__version__ = 1.1
warnings.filterwarnings('ignore')  # Turn off openpyxl Discarded range warnings


class XlsxTools:
    def __init__(self):
        self.xlsx_filename = ''
        self.wb = None
        self.ws = None

    def create_document(self, sheet_data: List[Dict], tab_name, xlsx_filename):
        """
        Write an xlsx file from [({'header1': 'dataA', 'header2':'dataB'}), ({'header1': 'dataC', 'header2':'dataD'})]
        :param sheet_data: list of dictionary entries for each row
        :param tab_name: Name of the tab
        :param xlsx_filename: path where file will be created
        """
        self.xlsx_filename = xlsx_filename
        self.wb = Workbook()
        self.ws = self.wb.active    # default: "Sheet"
        self.ws.title = tab_name
        self._write_data(sheet_data)
        self.format(sheet_data)
        self.wb.save(xlsx_filename)

    def _write_data(self, list_of_dict: List[Dict]):
        if not list_of_dict:
            return
        self.ws.append(list(list_of_dict[0].keys()))  # headers
        for row in list_of_dict:  # data
            self.ws.append(['%s' % data for data in list(row.values())])

    def add_work_sheet(self, sheet_data, tab_name, xlsx_filename=""):
        """
        Add worksheet. If xlsx_filename, open. If not, assume self.wb is loaded
        :param sheet_data: list of dictionaries to write to the sheet
        """
        if xlsx_filename:
            self.xlsx_filename = xlsx_filename
            self.wb = load_workbook(self.xlsx_filename)
        else:
            try:
                self.wb    #createDocument creates self.wb and should have been called
            except AttributeError:
                raise AttributeError("addWorkSheet: No filename provided, or didn't call createDocument first")
        self.ws = self.wb.create_sheet()
        self.ws.title = tab_name
        self._write_data(sheet_data)
        self.format(sheet_data)
        self.wb.save(self.xlsx_filename)

    def format(self, list_of_dict: List[Dict]):
        if not list_of_dict:
            return
        self.freeze_panes_first_row()
        self.auto_fit(list_of_dict)
        self.auto_filter()
        self.make_first_row_bold()

    def auto_fit(self, list_of_dict):
        """
        AutoFit Column Widths
        Sets the column widths automatically based on length. Assumes you've created self.wb / self.ws
        from http://stackoverflow.com/questions/13197574/python-openpyxl-column-width-size-adjust
        Ignore header length, find the max string length of each cell contents
        """
        column_widths = []
        for header_name in list(list_of_dict[0].keys()):
            # str() is needed in len(str()) as len(int) is an error
            column_widths.append(max(len(str(data[header_name])) if data[header_name] != None else 0 for data in list_of_dict))

        for i, column_width in enumerate(column_widths):
            if column_width < 10:
                column_width = 10
            if column_width > 50:
                column_width = 50
            self.ws.column_dimensions[get_column_letter(i+1)].width = column_width

    def header_row_reference(self):
        return "A1:" + get_column_letter(len(list(self.ws.rows)[0])) + "1"

    def auto_filter(self):
        self.ws.auto_filter.ref = self.header_row_reference()

    def freeze_panes_first_row_all(self):
        worksheets = self.wb.get_sheet_names()
        for worksheet in worksheets:
            self.wb[worksheet].freeze_panes = self.wb[worksheet]['A2']

    def freeze_panes_first_row(self):
        self.ws.freeze_panes = self.ws['A2']

    def make_first_row_bold(self):
        bold_style = NamedStyle(font=Font(bold=True), name='bold_style')
        for row in self.ws[self.header_row_reference()]:
            for cell in row:
                cell.style = bold_style

    @staticmethod
    def dict_reader(filename, tab_name, header_row_cell_value=''):
        """
        Creates list of ordered dictionaries of xlsx data (Recoded 4nov15)
        :param filename: xlsx filename to turn into dictionary
        :param tab_name: tab_name of the worksheet to read
        :param header_row_cell_value: the value of any cell indicating the header row
        :return: [{line1header1: value, ...},{line2header2: value}]
        """
        wb = load_workbook(filename=filename, read_only=True, data_only=True)
        ws = wb[tab_name]
        header = []
        result = []
        found = False if header_row_cell_value else True
        for r in ws.rows:
            if not found:
                for cell in r:
                    if cell.value == header_row_cell_value:
                        found = True
                        break
            if found:
                if not header:
                    header = [cell.value for cell in r]
                    continue
                result.append(OrderedDict(zip(header,
                                              [cell.value.strftime('%H:%M') if type(cell.value) is datetime.time
                                               else '00:00' if cell.value == datetime.datetime(1899, 12, 30, 0, 0)
                                               else cell.value.strftime('%d/%m/%Y') if type(cell.value) is datetime.datetime
                                               else '' if cell.value is None
                                               else str(cell.value)
                                               for cell in r])))
        return result

    def xlsx_to_csv(self, xlsx_to_read, csv_to_write, tab_name, delimeter=',', data_only=True, header_row_cell_value=''):
        # TODO: remove_no this method in re of using dict_reader + csv: dict_to_csv
        """
        Create csv from xlsx
        data_only=True: output value, not formula definition
        """
        wb = load_workbook(filename=xlsx_to_read, read_only=True, data_only=data_only)
        ws = wb[tab_name]
        with open(csv_to_write, 'w', newline='') as csvOut:
            cw = csv.writer(csvOut, delimiter=delimeter)
            found = False if header_row_cell_value else True
            for r in ws.rows:
                if not found:
                    for cell in r:
                        if cell.value == header_row_cell_value:
                            found = True
                            break
                if found:
                    cw.writerow([cell.value.strftime('%H:%M') if type(cell.value) is datetime.time
                                 else cell.value.strftime('%d/%m/%Y') if type(cell.value) is datetime.datetime
                                 else cell.value
                                 for cell in r])


        # data = self.dict_reader(xlsx_to_read, tab_name, header_row_cell_value=header_row_cell_value)
        # with open(csv_to_write, 'w', newline='') as csv_file:
        #     writer = csv.DictWriter(csv_file, list(data[0].keys()), delimiter=delimiter)
        #     writer.writeheader()
        #     for row in data:
        #         writer.writerow(row)